"""
core/excel_import.py — Import a legacy Master Deliverable List .xlsx.

Expected worksheet structure (per-set sheets, skip "Overall"):
  Row 1 (merged with row 2 for most cols):
    Col B: SET | DRAWING NUMBER | DRAWING DESCRIPTION | REV | DATE | ... | STATUS | NOTES
  Row 2: summary percentage labels on right side (skip)
  Row 3+: data rows

REV/DATE column pairs repeat (4-5 pairs observed in the wild).
Dates may be datetime objects or the literal string "-".

Returns:
  {
    "register": { ... },   # ready-to-save register dict
    "warnings": [...],
    "drawing_count": N
  }
"""

from __future__ import annotations

import re
from datetime import datetime, date
from typing import Any

import openpyxl

from core.register import _now_iso, VALID_STATUSES
from core.migration import migrate_register


_STATUS_NORMALISE = {
    "READY FOR SUBMITTAL": "READY FOR SUBMITTAL",
    "READY FOR DRAFTING": "READY FOR DRAFTING",
    "IN DESIGN": "IN DESIGN",
    "NOT CREATED YET": "NOT CREATED YET",
    # Common abbreviations / typos seen in the wild
    "RFS": "READY FOR SUBMITTAL",
    "RFD": "READY FOR DRAFTING",
    "IFD": "IN DESIGN",
    "NCY": "NOT CREATED YET",
}


def _cell_str(cell) -> str:
    """Return a clean string from a cell value."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _is_empty_or_dash(val: str) -> bool:
    return not val or val == "-" or val.lower() == "none"


def _find_header_col(ws, row_idx: int, header_name: str) -> int | None:
    """Search a row for a cell whose value matches header_name (case-insensitive).
    Returns the 1-based column index or None."""
    target = header_name.strip().upper()
    for col in range(1, ws.max_column + 1):
        v = _cell_str(ws.cell(row=row_idx, column=col)).upper()
        if v == target:
            return col
    return None


def _find_rev_date_columns(ws, header_row: int) -> list[tuple[int, int]]:
    """Return list of (rev_col, date_col) pairs by scanning the header row."""
    pairs: list[tuple[int, int]] = []
    last_rev_col: int | None = None
    for col in range(1, ws.max_column + 1):
        v = _cell_str(ws.cell(row=header_row, column=col)).upper()
        if v == "REV":
            last_rev_col = col
        elif v == "DATE" and last_rev_col is not None:
            pairs.append((last_rev_col, col))
            last_rev_col = None
    return pairs


def import_excel(path: str) -> dict[str, Any]:
    """Read a legacy Master Deliverable List .xlsx and return a register dict."""
    wb = openpyxl.load_workbook(path, data_only=True)

    warnings: list[str] = []
    sets: list[dict[str, Any]] = []
    total_drawings = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().upper() == "OVERALL":
            continue

        ws = wb[sheet_name]

        # ── Detect header row (row 1, possibly merged, so read merged values) ──
        # The merged cells in row 1 span row 2 as well; openpyxl reads the value
        # on the top-left cell of the merged range.  We search rows 1–3 for
        # "DRAWING NUMBER" to locate the header.
        header_row = None
        draw_num_col = None
        for r in (1, 2, 3):
            c = _find_header_col(ws, r, "DRAWING NUMBER")
            if c is not None:
                header_row = r
                draw_num_col = c
                break

        if header_row is None or draw_num_col is None:
            warnings.append(
                f"Sheet '{sheet_name}': could not find DRAWING NUMBER header — skipped."
            )
            continue

        desc_col = _find_header_col(ws, header_row, "DRAWING DESCRIPTION")
        status_col = _find_header_col(ws, header_row, "STATUS")
        notes_col = _find_header_col(ws, header_row, "NOTES")
        rev_date_pairs = _find_rev_date_columns(ws, header_row)

        # Data starts at header_row + 2 (skip the summary-percentage row 2)
        data_start = header_row + 2

        drawings: list[dict[str, Any]] = []

        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
            row_num = row[0].row

            draw_num = _cell_str(ws.cell(row=row_num, column=draw_num_col))
            if _is_empty_or_dash(draw_num):
                # Skip blank rows or summary rows
                continue

            description = ""
            if desc_col:
                description = _cell_str(ws.cell(row=row_num, column=desc_col))

            raw_status = ""
            if status_col:
                raw_status = _cell_str(ws.cell(row=row_num, column=status_col)).upper()
            status = _STATUS_NORMALISE.get(raw_status, "NOT CREATED YET")
            if raw_status and raw_status not in _STATUS_NORMALISE:
                warnings.append(
                    f"Sheet '{sheet_name}' row {row_num}: "
                    f"unknown status '{raw_status}' — defaulted to NOT CREATED YET."
                )

            notes_val = ""
            if notes_col:
                notes_val = _cell_str(ws.cell(row=row_num, column=notes_col))

            revisions: list[dict[str, str]] = []
            for rev_col, date_col in rev_date_pairs:
                rev_val = _cell_str(ws.cell(row=row_num, column=rev_col))
                date_val = _cell_str(ws.cell(row=row_num, column=date_col))
                if _is_empty_or_dash(rev_val):
                    continue
                # Normalise date: already string from _cell_str; ensure YYYY-MM-DD
                if not _is_empty_or_dash(date_val):
                    # Try to parse and reformat
                    date_val = _normalise_date(date_val)
                else:
                    date_val = ""
                revisions.append({"rev": rev_val, "date": date_val})

            drawings.append(
                {
                    "drawing_number": draw_num,
                    "description": description,
                    "status": status,
                    "notes": notes_val or None,
                    "revisions": revisions,
                }
            )
            total_drawings += 1

        sets.append({"name": sheet_name, "drawings": drawings})

    # Attempt to extract project number from the first drawing's number
    project_number = ""
    if sets and sets[0]["drawings"]:
        first_num = sets[0]["drawings"][0]["drawing_number"]
        m = re.match(r"(R3P-\d+)", first_num, re.IGNORECASE)
        if m:
            project_number = m.group(1).upper()

    # v1 here is intentional — this synthetic dict is the entry point for
    # migrate_register() which upgrades to current; never written to disk as v1
    register_v1: dict[str, Any] = {
        "schema_version": 1,
        "project_number": project_number,
        "project_name": "",
        "updated_at": _now_iso(),
        "sets": sets,
    }

    register = migrate_register(register_v1)

    return {
        "register": register,
        "warnings": warnings,
        "drawing_count": total_drawings,
    }


_DATE_FORMATS = [
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%d/%m/%Y",
    "%d-%b-%Y",
    "%d-%B-%Y",
    "%Y%m%d",
]


def _normalise_date(raw: str) -> str:
    """Try to parse raw date string and return ISO YYYY-MM-DD, or raw on failure."""
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw
