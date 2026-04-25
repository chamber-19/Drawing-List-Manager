"""
core/excel_export.py — Generate clean branded .xlsx output.

Function: export_full(path, register)
   Sheet 1: "Drawing Index"
     Always-visible cols A–G:
       Drawing Number | Description | Set | Phase | Current Rev | Current Date | Notes
     Grouped (collapsed) block H–Q:
       REV1 | DATE1 | REV2 | DATE2 | REV3 | DATE3 | REV4 | DATE4 | REV5 | DATE5

   Sheet 2: "Revision History"
     Flat audit log:
       Drawing Number | Rev | Phase | Percent | Date | Notes

   Branding: R3P palette, header fill, monospaced drawing numbers,
             frozen header row, autofilter.
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from core.register import current_display_rev


# ─── Colour palette (mirrors the JS design system tokens) ─────

_COPPER = "C4884D"
_BG_DARK = "252420"
_CREAM = "F0ECE4"
_MUTED = "A39E93"

# Status colours
_STATUS_STYLES: dict[str, dict[str, str]] = {
    "NOT CREATED YET": {"fill": "E8E6E0", "font": "4A4742"},
    "IN DESIGN": {"fill": "D6E4F0", "font": "1D4E6F"},
    "READY FOR DRAFTING": {"fill": "F5E9CC", "font": "7A5E12"},
    "READY FOR SUBMITTAL": {"fill": "D9ECD9", "font": "2C5F2C"},
}

_THIN = Side(style="thin", color="D0CCC5")
_THIN_BORDER = Border(
    left=_THIN, right=_THIN, top=_THIN, bottom=_THIN
)


# ─── Helpers ──────────────────────────────────────────────────

def _pf(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, color="F0ECE4", size=10, italic=False, name="Calibri") -> Font:
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _natural_sort_key(drawing_number: str):
    """Key for natural (human-friendly) sort of drawing numbers."""
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", drawing_number)
    ]


def _current_date(drawing: dict) -> str:
    """Return the date of the most recent revision, or empty string."""
    revs = drawing.get("revisions", [])
    if revs:
        return revs[-1].get("date", "")
    return ""


def _current_phase(drawing: dict) -> str:
    """Return the phase of the most recent revision, or empty string."""
    revs = drawing.get("revisions", [])
    if revs:
        return revs[-1].get("phase", "")
    return ""


# ─── Full Export ───────────────────────────────────────────────

def export_full(path: str, register: dict[str, Any]) -> None:
    """Write the full branded .xlsx register (two sheets)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    project_number = register.get("project_number", "")
    drawings_all = sorted(
        register.get("drawings", []),
        key=lambda d: _natural_sort_key(d.get("drawing_number", "")),
    )

    _write_index_sheet(wb, project_number, drawings_all)
    _write_history_sheet(wb, drawings_all)

    wb.save(path)


# ─── Sheet 1: Drawing Index ────────────────────────────────────

def _write_index_sheet(wb, project_number: str, drawings: list[dict]) -> None:
    ws = wb.create_sheet(title="Drawing Index")

    # Column definitions (always-visible A–G)
    always_cols = [
        ("Drawing Number", 22),
        ("Description",    48),
        ("Set",            10),
        ("Phase",           8),
        ("Current Rev",    16),
        ("Current Date",   14),
        ("Notes",          32),
    ]

    # Grouped cols H–Q: REV1 DATE1 … REV5 DATE5
    grouped_pairs = 5
    grouped_col_width = 8

    total_cols = len(always_cols) + grouped_pairs * 2

    for i, (_, width) in enumerate(always_cols):
        ws.column_dimensions[get_column_letter(i + 1)].width = width
    for j in range(grouped_pairs * 2):
        col_idx = len(always_cols) + j + 1
        ws.column_dimensions[get_column_letter(col_idx)].width = grouped_col_width

    # ── Row 1: Project header ──────────────────────────────────
    ws.row_dimensions[1].height = 28
    header_text = f"ROOT3POWER ENGINEERING  ·  {project_number}  ·  DRAWING INDEX"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=header_text)
    c.fill = _pf(_COPPER)
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.alignment = _align("center", "center")

    # ── Row 2: Column headers ──────────────────────────────────
    ws.row_dimensions[2].height = 18
    for i, (label, _) in enumerate(always_cols):
        c = ws.cell(row=2, column=i + 1, value=label)
        c.fill = _pf(_BG_DARK)
        c.font = Font(name="Calibri", bold=True, size=11, color=_CREAM)
        c.alignment = _align("center", "center")
        c.border = _THIN_BORDER

    # Grouped header labels
    for j in range(grouped_pairs):
        rev_col = len(always_cols) + j * 2 + 1
        date_col = rev_col + 1
        rev_num = j + 1
        for col, label in [(rev_col, f"REV{rev_num}"), (date_col, f"DATE{rev_num}")]:
            c = ws.cell(row=2, column=col, value=label)
            c.fill = _pf("3A3A35")  # slightly lighter than main header
            c.font = Font(name="Calibri", bold=True, size=10, color=_MUTED)
            c.alignment = _align("center", "center")
            c.border = _THIN_BORDER

    # Freeze at row 3, col A
    ws.freeze_panes = "A3"

    # Autofilter on always-visible columns
    ws.auto_filter.ref = f"A2:{get_column_letter(len(always_cols))}2"

    # Column grouping for H–Q (hidden by default)
    h_letter = get_column_letter(len(always_cols) + 1)
    q_letter = get_column_letter(total_cols)
    ws.column_dimensions.group(h_letter, q_letter, outline_level=1, hidden=True)
    ws.sheet_properties.outlinePr.summaryRight = False

    # ── Data rows ─────────────────────────────────────────────
    for idx, drawing in enumerate(drawings):
        row = idx + 3

        ws.row_dimensions[row].height = 20

        draw_num = drawing.get("drawing_number", "")
        description = drawing.get("description", "")
        set_val = drawing.get("set", "")
        phase = _current_phase(drawing)
        cur_rev = current_display_rev(drawing)
        cur_date = _current_date(drawing)
        notes = drawing.get("notes") or "—"
        revisions = drawing.get("revisions", [])

        row_data = [draw_num, description, set_val, phase, cur_rev, cur_date, notes]

        for col_idx, value in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=col_idx, value=value)
            if col_idx == 1:
                c.font = Font(name="Courier New", bold=True, size=10, color=_COPPER)
                c.alignment = _align("left", "center")
            elif col_idx == 7:
                is_placeholder = notes == "—"
                c.font = Font(
                    name="Calibri", size=9, italic=True,
                    color=_MUTED if is_placeholder else _CREAM,
                )
                c.alignment = _align("left", "center", wrap=True)
            else:
                c.font = Font(name="Calibri", size=10, color=_CREAM)
                c.alignment = _align("left", "center")
            c.border = _THIN_BORDER

        # Grouped revision cols
        for j in range(grouped_pairs):
            rev_col = len(always_cols) + j * 2 + 1
            date_col = rev_col + 1
            if j < len(revisions):
                rev_entry = revisions[j]
                rev_val = rev_entry.get("rev", "")
                date_val = rev_entry.get("date", "")
            else:
                rev_val = ""
                date_val = ""
            for col, val in [(rev_col, rev_val), (date_col, date_val)]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name="Courier New", size=9, color=_MUTED)
                c.alignment = _align("center", "center")
                c.border = _THIN_BORDER

        # Zebra striping
        if idx % 2 == 1:
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=row, column=col)
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "000000"):
                    cell.fill = _pf("2A2924")

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"


# ─── Sheet 2: Revision History ────────────────────────────────

def _write_history_sheet(wb, drawings: list[dict]) -> None:
    ws = wb.create_sheet(title="Revision History")

    col_defs = [
        ("Drawing Number", 22),
        ("Rev",             6),
        ("Phase",           8),
        ("Percent",        10),
        ("Date",           12),
        ("Notes",          40),
    ]
    for i, (_, width) in enumerate(col_defs):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    # ── Row 1: Header ──────────────────────────────────────────
    ws.row_dimensions[1].height = 18
    for i, (label, _) in enumerate(col_defs):
        c = ws.cell(row=1, column=i + 1, value=label)
        c.fill = _pf(_BG_DARK)
        c.font = Font(name="Calibri", bold=True, size=11, color=_CREAM)
        c.alignment = _align("center", "center")
        c.border = _THIN_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_defs))}1"

    row = 2
    for drawing in drawings:
        draw_num = drawing.get("drawing_number", "")
        draw_notes = drawing.get("notes") or ""
        for rev_entry in drawing.get("revisions", []):
            ws.row_dimensions[row].height = 18
            rev = rev_entry.get("rev", "")
            phase = rev_entry.get("phase", "")
            percent = rev_entry.get("percent")
            date = rev_entry.get("date", "")

            row_vals = [draw_num, rev, phase, percent, date, draw_notes]
            for col_idx, value in enumerate(row_vals, start=1):
                c = ws.cell(row=row, column=col_idx, value=value)
                if col_idx == 1:
                    c.font = Font(name="Courier New", bold=True, size=10, color=_COPPER)
                    c.alignment = _align("left", "center")
                else:
                    c.font = Font(name="Calibri", size=9, color=_CREAM)
                    c.alignment = _align("center", "center")
                c.border = _THIN_BORDER
            row += 1

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
